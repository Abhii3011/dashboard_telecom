# dashboard_refactored.py
import streamlit as st
import plotly.express as px
import pandas as pd
import warnings
import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment

warnings.filterwarnings("ignore")

# ---------------------
# Page config & CSS
# ---------------------
st.set_page_config(page_title="Telecom Network Dashboard", layout="wide")

st.markdown(
    """
    <style>
        .main-title {
            font-size: 28px !important;
            font-weight: bold;
            text-align: center;
            color: #ffffff;
            padding: 10px;
            background-color: #003366;
            border-bottom: 3px solid #0055A4;
        }
        .chart-header {
            display:flex;
            justify-content:space-between;
            align-items:center;
            font-size:20px;
            font-weight:bold;
            margin-top:25px;
            padding:8px;
            background-color:#f0f2f6;
            border-radius:6px;
        }
        .legend-box {display:flex; gap:12px; align-items:center;}
    </style>
    """,
    unsafe_allow_html=True,
)

st.markdown('<div class="main-title">Telecom Network Performance Dashboard</div>', unsafe_allow_html=True)

# ---------------------
# Data loading
# ---------------------
@st.cache_data(ttl=300)
def load_csvs(perc_path="sample_dataset.csv", delay_path="delay_dataset.csv"):
    try:
        df_perc = pd.read_csv(perc_path)
        df_delay = pd.read_csv(delay_path)
    except FileNotFoundError:
        st.error("CSV files not found. Make sure 'sample_dataset.csv' and 'delay_dataset.csv' exist in the project folder.")
        st.stop()

    def normalize_cols(df):
        df = df.copy()
        df.columns = [
            c.strip().lower().replace("/", "_").replace(" ", "_") for c in df.columns
        ]
        if "filedate" in df.columns:
            df["filedate"] = pd.to_datetime(df["filedate"], errors="coerce")
        if "enodeb_gnodeb" in df.columns:
            df.rename(columns={"enodeb_gnodeb": "gNodeB"}, inplace=True)
        df["market"] = df["market"].astype(str) if "market" in df.columns else df.get("market", "")
        return df

    df_perc = normalize_cols(df_perc)
    df_delay = normalize_cols(df_delay)
    return df_perc, df_delay

df_perc, df_delay = load_csvs()
metadata_cols = {"region", "market", "gNodeB", "filedate", "risk"}
original_interval_cols = [c for c in df_perc.columns if c not in metadata_cols]

HOUR_MAP = {f"{h:02d}": [f"{h:02d}:{m:02d}" for m in range(0, 60, 15)] for h in range(24)}

def resample_to_hourly(df, interval_cols):
    df = df.copy()
    for hour, cols in HOUR_MAP.items():
        existing = [c for c in cols if c in df.columns]
        if existing:
            df[hour] = df[existing].mean(axis=1)
    cols = ["region", "market", "gNodeB", "filedate"] + list(HOUR_MAP.keys())
    present = [c for c in cols if c in df.columns]
    return df[present]

def resample_to_daily(df, interval_cols):
    df = df.copy()
    existing_intervals = [c for c in interval_cols if c in df.columns]
    if existing_intervals:
        df["daily_avg"] = df[existing_intervals].mean(axis=1)
    else:
        df["daily_avg"] = pd.NA
    df["risk"] = df.get("risk", 0)
    return df[["region", "market", "gNodeB", "filedate", "daily_avg", "risk"]]

def df_to_excel_with_colors(df, mode="arrival"):
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    header = ["" if df.index.name is None else df.index.name] + list(df.columns)
    ws.append(header)
    for idx, row in df.reset_index().iterrows():
        ws.append(list(row))
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=ws.max_column):
        for cell in row:
            try:
                val = float(cell.value)
            except Exception:
                continue
            fill_color = None
            font_color = "000000"
            if mode == "arrival":
                if val == 100:
                    fill_color, font_color = "08306b", "FFFFFF"
                elif val >= 75:
                    fill_color, font_color = "1f78ff", "FFFFFF"
                elif val >= 50:
                    fill_color = "73b3ff"
                elif val >= 25:
                    fill_color = "d0e7ff"
                else:
                    fill_color, font_color = "FF0000", "FFFFFF"
            else:
                if val <= 5:
                    fill_color = "d0e7ff"
                elif val <= 10:
                    fill_color = "73b3ff"
                elif val <= 15:
                    fill_color, font_color = "1f78ff", "FFFFFF"
                else:
                    fill_color, font_color = "08306b", "FFFFFF"
            if fill_color:
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            cell.font = Font(color=font_color)
            cell.alignment = Alignment(horizontal="center", vertical="center")
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()

def apply_heatmap_style(df, mode="arrival", fmt=None):
    def style_func(v):
        if pd.isna(v):
            return "background-color:white; color:black"
        try:
            val = float(v)
        except Exception:
            return ""
        if mode == "arrival":
            if val == 100:
                return "background-color:#08306b; color:white"
            elif val >= 75:
                return "background-color:#1f78ff; color:white"
            elif val >= 50:
                return "background-color:#73b3ff"
            elif val >= 25:
                return "background-color:#d0e7ff"
            else:
                return "background-color:#FF0000; color:white"
        else:
            if val <= 5:
                return "background-color:#d0e7ff"
            elif val <= 10:
                return "background-color:#73b3ff"
            elif val <= 15:
                return "background-color:#1f78ff; color:white"
            else:
                return "background-color:#08306b; color:white"
    styler = df.style.applymap(style_func)
    if fmt:
        styler = styler.format(fmt)
    return styler

def plot_bar(df, x, y, title, hline=None, text_auto=True):
    fig = px.bar(df, x=x, y=y, title=title, text_auto=text_auto)
    fig.update_xaxes(categoryorder="category ascending")
    if hline is not None:
        fig.add_hline(y=hline["y"], line_dash="dash", line_color=hline.get("color", "red"),
                      annotation_text=hline.get("text", "Threshold"))
    return fig

# ---------------------
# Sidebar filters & options (Cascading)
# ---------------------
st.sidebar.header("Filters")

combined = pd.concat(
    [
        df_perc[["region", "market", "gNodeB", "filedate"]],
        df_delay[["region", "market", "gNodeB", "filedate"]],
    ],
    ignore_index=True,
)

view_mode = st.sidebar.radio("View Mode", ["15-min Intervals", "Hourly", "Daily"])

# Build data views depending on selected mode
if view_mode == "Hourly":
    delay_data = resample_to_hourly(df_delay, original_interval_cols)
    sample_data = resample_to_hourly(df_perc, original_interval_cols)
    interval_cols = [f"{h:02d}" for h in range(24)]
elif view_mode == "Daily":
    delay_daily = resample_to_daily(df_delay, original_interval_cols)
    sample_daily = resample_to_daily(df_perc, original_interval_cols)
    if "filedate" in delay_daily.columns:
        delay_daily["date"] = pd.to_datetime(delay_daily["filedate"]).dt.normalize()
    else:
        delay_daily["date"] = pd.NaT
    if "filedate" in sample_daily.columns:
        sample_daily["date"] = pd.to_datetime(sample_daily["filedate"]).dt.normalize()
    else:
        sample_daily["date"] = pd.NaT
    delay_daily = (
        delay_daily.groupby(["region", "market", "gNodeB", "date"], as_index=False)
        .agg({"daily_avg": "mean"})
    )
    sample_daily = (
        sample_daily.groupby(["region", "market", "gNodeB", "date"], as_index=False)
        .agg({"daily_avg": "mean", "risk": "sum"})
    )
    combos = combined[["region", "market", "gNodeB"]].dropna().drop_duplicates().reset_index(drop=True)
    last_date = pd.to_datetime(combined["filedate"]).max().normalize()
    last5_days = pd.date_range(end=last_date, periods=5, freq="D")
    dates_df = pd.DataFrame({"date": last5_days})
    combos["_key"] = 1
    dates_df["_key"] = 1
    grid = combos.merge(dates_df, on="_key").drop(columns=["_key"])
    sample_data = grid.merge(sample_daily, on=["region", "market", "gNodeB", "date"], how="left")
    delay_data = grid.merge(delay_daily, on=["region", "market", "gNodeB", "date"], how="left")
    interval_cols = list(last5_days)
else:
    delay_data = df_delay.copy()
    sample_data = df_perc.copy()
    interval_cols = original_interval_cols

# --- Cascading Filters ---
risk_prone_checkbox = st.sidebar.checkbox("Show Problematic Sites Only")

# 1. Region
all_regions = sorted(combined["region"].dropna().unique().tolist())
selected_region = st.sidebar.selectbox("Select Region", ["All"] + all_regions)
if selected_region == "All":
    filtered_markets = sorted(combined["market"].dropna().unique().tolist())
else:
    filtered_markets = sorted(combined[combined["region"] == selected_region]["market"].dropna().unique().tolist())

# 2. Market
selected_market = st.sidebar.selectbox("Select Market", ["All"] + filtered_markets)
if selected_market == "All":
    if selected_region == "All":
        filtered_dates = combined["filedate"].dropna().dt.date.unique()
    else:
        filtered_dates = combined[combined["region"] == selected_region]["filedate"].dropna().dt.date.unique()
else:
    if selected_region == "All":
        filtered_dates = combined[combined["market"] == selected_market]["filedate"].dropna().dt.date.unique()
    else:
        filtered_dates = combined[(combined["region"] == selected_region) & (combined["market"] == selected_market)]["filedate"].dropna().dt.date.unique()
filtered_dates = sorted(filtered_dates)

# 3. Date
selected_date = st.sidebar.selectbox("Select Date", ["All"] + [str(d) for d in filtered_dates])
if selected_date == "All":
    if selected_market == "All":
        if selected_region == "All":
            filtered_gnodes = combined["gNodeB"].dropna().unique()
        else:
            filtered_gnodes = combined[combined["region"] == selected_region]["gNodeB"].dropna().unique()
    else:
        if selected_region == "All":
            filtered_gnodes = combined[combined["market"] == selected_market]["gNodeB"].dropna().unique()
        else:
            filtered_gnodes = combined[(combined["region"] == selected_region) & (combined["market"] == selected_market)]["gNodeB"].dropna().unique()
else:
    date_dt = pd.to_datetime(selected_date)
    mask = True
    if selected_region != "All":
        mask = mask & (combined["region"] == selected_region)
    if selected_market != "All":
        mask = mask & (combined["market"] == selected_market)
    mask = mask & (combined["filedate"].dt.date == date_dt.date())
    filtered_gnodes = combined[mask]["gNodeB"].dropna().unique()
filtered_gnodes = sorted(filtered_gnodes)

# 4. gNodeB
selected_gnodes = st.sidebar.multiselect("Select gNodeB(s)", ["All"] + filtered_gnodes, default=["All"])
if "All" in selected_gnodes or not selected_gnodes:
    selected_gnodes = filtered_gnodes

# --- Problematic Sites Only ---
if risk_prone_checkbox:
    if view_mode == "Daily":
        file_mask = sample_data["daily_avg"].fillna(100) < 100
        delay_mask = delay_data["daily_avg"].fillna(0) > 20
        file_risk_sites = sample_data.loc[file_mask, "gNodeB"].unique() if "gNodeB" in sample_data else []
        delay_risk_sites = delay_data.loc[delay_mask, "gNodeB"].unique() if "gNodeB" in delay_data else []
    else:
        sample_interval_cols = [c for c in interval_cols if c in sample_data.columns]
        delay_interval_cols = [c for c in interval_cols if c in delay_data.columns]
        if sample_interval_cols:
            file_risk_sites = sample_data[sample_interval_cols].min(axis=1) < 100
            file_risk_sites = sample_data.loc[file_risk_sites, "gNodeB"].unique()
        else:
            file_risk_sites = []
        if delay_interval_cols:
            delay_risk_sites = delay_data[delay_interval_cols].max(axis=1) > 20
            delay_risk_sites = delay_data.loc[delay_risk_sites, "gNodeB"].unique()
        else:
            delay_risk_sites = []
    risk_gnodes = sorted(list(set(list(file_risk_sites) + list(delay_risk_sites))))
    if not risk_gnodes:
        st.info("No problematic sites found with the current filters.")
    else:
        selected_gnodes = risk_gnodes

# --- Filtered Data ---
def apply_filters(df, is_daily=False):
    df = df.copy()
    if selected_region != "All" and "region" in df.columns:
        df = df[df["region"] == selected_region]
    if selected_market != "All" and "market" in df.columns:
        df = df[df["market"] == selected_market]
    if selected_date != "All":
        if is_daily:
            if "date" in df.columns:
                df = df[df["date"] == pd.to_datetime(selected_date)]
        else:
            if "filedate" in df.columns:
                df = df[df["filedate"].dt.date == pd.to_datetime(selected_date).date()]
    if "gNodeB" in df.columns and selected_gnodes:
        df = df[df["gNodeB"].isin(selected_gnodes)]
    return df

if view_mode == "Daily":
    delay_filtered = apply_filters(delay_data, is_daily=True)
    sample_filtered = apply_filters(sample_data, is_daily=True)
else:
    delay_filtered = apply_filters(delay_data, is_daily=False)
    sample_filtered = apply_filters(sample_data, is_daily=False)

total_gnodes = len(delay_filtered["gNodeB"].unique()) if "gNodeB" in delay_filtered.columns else 0

# ---------------------
# UI: 1) File Arrival Track
# ---------------------
st.markdown(f'<div class="chart-header">File Arrival Track ({view_mode}) <span>Total gNodeBs: {total_gnodes}</span></div>', unsafe_allow_html=True)
if sample_filtered.empty:
    st.info("No data available for selected filters (File Arrival).")
else:
    if view_mode == "Daily":
        pivot_table = sample_filtered.pivot_table(index="gNodeB", columns="date", values="daily_avg", aggfunc="mean")
        pivot_table = pivot_table.reindex(columns=sorted(pivot_table.columns))
        styler = apply_heatmap_style(pivot_table, mode="arrival", fmt="{:.0f}%")
        st.dataframe(styler, width="stretch")
        excel_bytes = df_to_excel_with_colors(pivot_table, mode="arrival")
        st.download_button("Download File Arrival Excel", data=excel_bytes, file_name="file_arrival.xlsx")
    else:
        present_intervals = [c for c in interval_cols if c in sample_filtered.columns]
        if not present_intervals:
            st.info("No interval columns present for File Arrival view.")
        else:
            pivot_table = sample_filtered.pivot_table(index="gNodeB", values=present_intervals, aggfunc="mean")
            styler = apply_heatmap_style(pivot_table, mode="arrival", fmt="{:.0f}%")
            st.dataframe(styler, width="stretch")
            excel_bytes = df_to_excel_with_colors(pivot_table, mode="arrival")
            st.download_button("Download File Arrival Excel", data=excel_bytes, file_name="file_arrival.xlsx")
    st.markdown(
        """
        **Legend:** <div style="display:flex;gap:15px;">
            <div style="background-color:#08306b;width:20px;height:20px;border:1px solid #000"></div> 100% (Perfect)
            <div style="background-color:#1f78ff;width:20px;height:20px;border:1px solid #000"></div> ≥75%
            <div style="background-color:#73b3ff;width:20px;height:20px;border:1px solid #000"></div> ≥50%
            <div style="background-color:#d0e7ff;width:20px;height:20px;border:1px solid #000"></div> ≥25%
            <div style="background-color:#FF0000;width:20px;height:20px;border:1px solid #000"></div> <25%
        </div>
        """,
        unsafe_allow_html=True,
    )

# ---------------------
# UI: 2) Delay Heatmap
# ---------------------
st.markdown(f'<div class="chart-header">Delay Heatmap ({view_mode}) <span>Total gNodeBs: {total_gnodes}</span></div>', unsafe_allow_html=True)
if delay_filtered.empty:
    st.info("No data available for selected filters (Delay Heatmap).")
else:
    if view_mode == "Daily":
        pivot_delay = delay_filtered.pivot_table(index="gNodeB", columns="date", values="daily_avg", aggfunc="mean")
        pivot_delay = pivot_delay.reindex(columns=sorted(pivot_delay.columns))
        styled_delay = apply_heatmap_style(pivot_delay, mode="delay", fmt="{:.1f} min")
        st.dataframe(styled_delay, width="stretch")
        excel_bytes = df_to_excel_with_colors(pivot_delay, mode="delay")
        st.download_button("Download Delay Heatmap Excel", data=excel_bytes, file_name="delay_heatmap.xlsx")
    else:
        present_intervals = [c for c in interval_cols if c in delay_filtered.columns]
        if not present_intervals:
            st.info("No interval columns present for Delay Heatmap view.")
        else:
            pivot_delay = delay_filtered.pivot_table(index="gNodeB", values=present_intervals, aggfunc="mean")
            styled_delay = apply_heatmap_style(pivot_delay, mode="delay", fmt="{:.1f} min")
            st.dataframe(styled_delay, width="stretch")
            excel_bytes = df_to_excel_with_colors(pivot_delay, mode="delay")
            st.download_button("Download Delay Heatmap Excel", data=excel_bytes, file_name="delay_heatmap.xlsx")
    st.markdown(
        """
        **Legend (minutes):** <div style="display:flex;gap:15px;">
            <div style="background-color:#d0e7ff;width:20px;height:20px;border:1px solid #000"></div> 0–5 min
            <div style="background-color:#73b3ff;width:20px;height:20px;border:1px solid #000"></div> 5–10 min
            <div style="background-color:#1f78ff;width:20px;height:20px;border:1px solid #000"></div> 10–15 min
            <div style="background-color:#08306b;width:20px;height:20px;border:1px solid #000"></div> 15+ min
        </div>
        """,
        unsafe_allow_html=True,
    )

# ---------------------
# UI: 3) Latency Trend
# ---------------------
st.markdown(f'<div class="chart-header">📊 Latency Trend by gNodeB ({view_mode}) <span>Total gNodeBs: {total_gnodes}</span></div>', unsafe_allow_html=True)
if delay_filtered.empty:
    st.info("No data available for Latency Trend.")
else:
    gnodeb_options = sorted(delay_filtered["gNodeB"].dropna().unique().tolist())
    if not gnodeb_options:
        st.info("No gNodeB data present for Latency Trend.")
    else:
        gnodeb_choice = st.selectbox("Select gNodeB for Latency Trend", options=gnodeb_options, index=0)
        if view_mode == "Daily":
            melted_delay = delay_filtered[delay_filtered["gNodeB"] == gnodeb_choice][["date", "daily_avg"]].copy()
            if melted_delay.empty:
                st.info("No daily latency data for selected gNodeB.")
            else:
                melted_delay = melted_delay.sort_values("date")
                melted_delay["date_str"] = melted_delay["date"].dt.strftime("%Y-%m-%d")
                fig_delay = plot_bar(melted_delay, x="date_str", y="daily_avg", title=f"Latency Trend - gNodeB: {gnodeb_choice}",)
                fig_delay.add_hline(y=10, line_dash="dash", line_color="red", annotation_text="Threshold = 10 min")
                st.plotly_chart(fig_delay, use_container_width=True)
        else:
            present_intervals = [c for c in interval_cols if c in delay_filtered.columns]
            subset = delay_filtered[delay_filtered["gNodeB"] == gnodeb_choice]
            if subset.empty or not present_intervals:
                st.info("No interval data to plot Latency Trend.")
            else:
                melted_delay = subset.melt(id_vars=["gNodeB"], value_vars=present_intervals, var_name="interval", value_name="delay")
                fig_delay = px.bar(melted_delay, x="interval", y="delay", color="gNodeB", barmode="group", title=f"Latency Trend - gNodeB: {gnodeb_choice}")
                fig_delay.add_hline(y=10, line_dash="dash", line_color="red", annotation_text="Threshold = 10 min")
                fig_delay.update_xaxes(categoryorder="category ascending")
                st.plotly_chart(fig_delay, use_container_width=True)

# ---------------------
# UI: 4) Average Delay by Region & Market
# ---------------------
st.markdown(f'<div class="chart-header">📍 Average Delay by Region & Market ({view_mode}) <span>Total gNodeBs: {total_gnodes}</span></div>', unsafe_allow_html=True)
if delay_filtered.empty:
    st.info("No data available for Average Delay charts.")
else:
    col3, col4 = st.columns(2)
    if view_mode == "Daily":
        delay_avg_region = delay_filtered.groupby("region")["daily_avg"].mean().reset_index(name="avg_delay")
        delay_avg_market = delay_filtered.groupby("market")["daily_avg"].mean().reset_index(name="avg_delay")
        fig3 = plot_bar(delay_avg_region, x="region", y="avg_delay", title="Average Delay by Region",)
        col3.plotly_chart(fig3, use_container_width=True)
        fig4 = plot_bar(delay_avg_market, x="market", y="avg_delay", title="Average Delay by Market",)
        col4.plotly_chart(fig4, use_container_width=True)
    else:
        present_intervals = [c for c in interval_cols if c in delay_filtered.columns]
        if not present_intervals:
            st.info("No interval columns to compute average delay.")
        else:
            delay_avg_region = delay_filtered.groupby("region")[present_intervals].mean().mean(axis=1).reset_index(name="avg_delay")
            delay_avg_market = delay_filtered.groupby("market")[present_intervals].mean().mean(axis=1).reset_index(name="avg_delay")
            fig3 = plot_bar(delay_avg_region, x="region", y="avg_delay", title="Average Delay by Region")
            col3.plotly_chart(fig3, use_container_width=True)
            fig4 = plot_bar(delay_avg_market, x="market", y="avg_delay", title="Average Delay by Market")
            col4.plotly_chart(fig4, use_container_width=True)

# ---------------------
# UI: 5) Problematic Sites Count
# ---------------------
st.markdown(f'<div class="chart-header">⚠️ Problematic Sites Count ({view_mode}) <span>Total gNodeBs: {total_gnodes}</span></div>', unsafe_allow_html=True)
if sample_filtered.empty or "risk" not in sample_filtered.columns:
    st.info("No 'risk' column present or no sample data available for Problematic Sites Count.")
else:
    col5, col6 = st.columns(2)
    risk_region = sample_filtered.groupby("region")["risk"].sum().reset_index()
    fig5 = plot_bar(risk_region, x="region", y="risk", title="Problematic Sites by Region")
    col5.plotly_chart(fig5, use_container_width=True)
    risk_market = sample_filtered.groupby("market")["risk"].sum().reset_index()
    fig6 = plot_bar(risk_market, x="market", y="risk", title="Problematic Sites by Market")
    col6.plotly_chart(fig6, use_container_width=True)

# ---------------------
# UI: 6) Market Pie (Zero-Interval distribution)
# ---------------------
st.markdown(f'<div class="chart-header">🥧 Zero-Interval Distribution by Market ({view_mode}) <span>Total gNodeBs: {total_gnodes}</span></div>', unsafe_allow_html=True)
if sample_filtered.empty:
    st.info("No sample data to create Zero-Interval distribution.")
else:
    if view_mode != "Daily":
        present_intervals = [c for c in interval_cols if c in sample_filtered.columns]
        if present_intervals:
            melted_percent = sample_filtered.melt(id_vars=["market"], value_vars=present_intervals, var_name="interval", value_name="value")
            melted_percent["zero_flag"] = melted_percent["value"].apply(lambda x: 1 if x == 0 else 0)
            pie_df = melted_percent.groupby("market")["zero_flag"].sum().reset_index()
            fig7 = px.pie(pie_df, values="zero_flag", names="market", title="Zero-Interval Counts by Market")
            st.plotly_chart(fig7, use_container_width=True)
        else:
            st.info("No interval columns available to compute Zero-Interval distribution.")
    else:
        if "daily_avg" in sample_filtered.columns:
            sample_filtered["zero_flag"] = sample_filtered["daily_avg"].apply(lambda x: 1 if pd.notna(x) and x == 0 else 0)
            pie_df = sample_filtered.groupby("market")["zero_flag"].sum().reset_index()
            fig7 = px.pie(pie_df, values="zero_flag", names="market", title="Zero-Interval Counts by Market (last 5 days)")
            st.plotly_chart(fig7, use_container_width=True)
        else:
            st.info("No daily averages to create Zero-Interval pie for Daily view.")